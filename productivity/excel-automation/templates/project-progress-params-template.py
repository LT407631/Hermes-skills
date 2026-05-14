"""
Project Progress Table with Parameter-Driven Conditional Formatting
========================================================================
Use as a starting template. Adapt column count, rule thresholds,
role names, and sheet names to your specific workflow.

Usage:
  python project-progress-params-template.py

Produces: project-progress.xlsx with 3 sheets:
  1. 参数调节 — user-adjustable yellow/red day thresholds
  2. Sheet2 — workflow step definitions (optional)
  3. 项目进度 — main tracking table with auto-warning conditional formatting

Author: Generated for 腾哥's 全屋定制工厂 project management
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.formatting.formatting import ConditionalFormattingList
from datetime import datetime

# ── CONFIG ──────────────────────────────────────────────────────────────────
OUTPUT_PATH = '/mnt/c/Users/LT-PC/Desktop/project-progress.xlsx'

# Define columns: (label, width)
HEADERS = [
    ('客单号', 12), ('销售经理', 12), ('经销商', 18), ('终端地址', 24),
    ('设计师', 10), ('拆单', 10),
    ('接单时间', 14),       # ← baseline column (col G / col 7)
    ('报价时间', 14),
    ('回款时间', 14),
    ('采购时间', 14),
    ('图纸时间', 14),
    ('拆单完成', 14),
    ('采购到厂', 14),
    ('生产开始', 14),
    ('入库完成', 14),
    ('合同金额', 14), ('投影面积', 12),
]

# Parameter sheet columns
PARAM_HEADERS = ['环节', '变黄天数', '变红天数', '说明']

# (环节名称, 黄天数, 红天数, 说明, 对应进度表列字母)
PARAMS = [
    ('报价时间',      1,  2,  '从接单日算起，超X天预警/超期', 'H'),
    ('回款时间',      1,  2,  '从报价日算起',                'I'),
    ('采购时间',      1,  2,  '从回款日算起',                'J'),
    ('图纸时间',      3,  4,  '从接单日算起',                'K'),
    ('拆单完成',      7,  8,  '从接单日算起',                'L'),
    ('采购到厂',      7,  8,  '从接单日算起',                'M'),
    ('生产开始',      8,  9,  '从接单日算起',                'N'),
    ('入库完成',     15, 16,  '从接单日算起',                'O'),
]

# Sample data row (empty dates to demonstrate warning colors)
SAMPLE_DATA = [
    'XD-2026-001',          # 客单号
    '张三',                  # 销售经理
    '天工全屋定制',           # 经销商
    '洛阳市洛龙区XX小区',     # 终端地址
    '李四',                  # 设计师
    '王五',                  # 拆单
    datetime(2026, 5, 10),   # 接单时间 (2 days ago → triggers warnings)
    None,                    # 报价时间 (empty → warning)
    None,                    # 回款时间
    None,                    # 采购时间
    None,                    # 图纸时间
    None,                    # 拆单完成
    None,                    # 采购到厂
    None,                    # 生产开始
    None,                    # 入库完成
    58000,                   # 合同金额
    86,                      # 投影面积
]

BASELINE_COL = 'G'   # Column letter for the baseline date (接单时间)
DATA_START_ROW = 2   # First data row
DATA_END_ROW = 100   # Last row for conditional formatting range

# ── STYLES ──────────────────────────────────────────────────────────────────
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

param_title_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
param_title_font = Font(name='微软雅黑', bold=True, size=14, color='FFFFFF')

yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')

# ── BUILD ───────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── Sheet 1: 参数调节 ──
ws_param = wb.active
ws_param.title = '参数调节'

# Title row
ws_param.merge_cells('A1:D1')
c = ws_param['A1']
c.value = '参数调节表'
c.font = param_title_font
c.fill = param_title_fill
c.alignment = center_align
ws_param.row_dimensions[1].height = 36

# Headers
for col, h in enumerate(PARAM_HEADERS, 1):
    c = ws_param.cell(row=2, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align
    c.border = thin_border

# Data rows (start from row 3)
for row_idx, (name, yd, rd, note) in enumerate(PARAMS, 3):
    ws_param.cell(row=row_idx, column=1, value=name).border = thin_border
    ws_param.cell(row=row_idx, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws_param.cell(row=row_idx, column=1).font = Font(name='微软雅黑', size=11)

    cy = ws_param.cell(row=row_idx, column=2, value=yd)
    cy.border = thin_border; cy.alignment = center_align
    cy.font = Font(name='微软雅黑', size=11, bold=True, color='CC8800')

    cr = ws_param.cell(row=row_idx, column=3, value=rd)
    cr.border = thin_border; cr.alignment = center_align
    cr.font = Font(name='微软雅黑', size=11, bold=True, color='CC0000')

    cn = ws_param.cell(row=row_idx, column=4, value=note)
    cn.border = thin_border; cn.alignment = Alignment(horizontal='left', vertical='center')
    cn.font = Font(name='微软雅黑', size=9, color='666666')

# Column widths
ws_param.column_dimensions['A'].width = 16
ws_param.column_dimensions['B'].width = 14
ws_param.column_dimensions['C'].width = 14
ws_param.column_dimensions['D'].width = 36

# Usage tips
tip_row = 3 + len(PARAMS) + 1
tips = [
    '💡 使用说明',
    '1. 直接修改"变黄天数"和"变红天数"的数字即可调整预警规则',
    '2. 修改后保存文件，项目进度表的条件格式会自动更新',
    '3. 如果修改后颜色未更新，关闭文件重新打开即可',
]
for i, tip in enumerate(tips):
    r = tip_row + i
    ws_param.merge_cells(f'A{r}:D{r}')
    c = ws_param[f'A{r}']
    c.value = tip
    c.font = Font(name='微软雅黑', bold=(i==0), size=11 if i==0 else 10,
                  color='2F5496' if i==0 else '666666')

# ── Sheet 2: 项目进度 ──
ws = wb.create_sheet('项目进度')

# Headers
for col_idx, (label, width) in enumerate(HEADERS, 1):
    c = ws.cell(row=1, column=col_idx, value=label)
    c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = thin_border
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 28

# Sample data
for col_idx, val in enumerate(SAMPLE_DATA, 1):
    c = ws.cell(row=2, column=col_idx, value=val)
    c.border = thin_border; c.alignment = center_align
    if isinstance(val, datetime):
        c.number_format = 'YYYY-MM-DD'
    elif isinstance(val, (int, float)):
        c.number_format = '#,##0'

ws.row_dimensions[2].height = 24

# ── Conditional Formatting: Cross-sheet references ──
ws.conditional_formatting = ConditionalFormattingList()

for (name, yd, rd, note, col_letter) in PARAMS:
    param_row = 3 + [p[0] for p in PARAMS].index(name)  # row in 参数调节 sheet

    # Yellow warning
    ws.conditional_formatting.add(
        f'{col_letter}{DATA_START_ROW}:{col_letter}{DATA_END_ROW}',
        FormulaRule(
            formula=[f'AND({col_letter}{DATA_START_ROW}="",TODAY()-${BASELINE_COL}{DATA_START_ROW}>=参数调节!$B${param_row})'],
            fill=yellow_fill
        )
    )
    # Red overdue (added second so it takes priority)
    ws.conditional_formatting.add(
        f'{col_letter}{DATA_START_ROW}:{col_letter}{DATA_END_ROW}',
        FormulaRule(
            formula=[f'AND({col_letter}{DATA_START_ROW}="",TODAY()-${BASELINE_COL}{DATA_START_ROW}>=参数调节!$C${param_row})'],
            fill=red_fill
        )
    )

# ── Save ──
wb.save(OUTPUT_PATH)
print(f'✅ 已生成: {OUTPUT_PATH}')
print(f'   工作表: {wb.sheetnames}')
print(f'   参数: 3+ 行可调节, {len(PARAMS)} 个环节的条件格式已设置')
